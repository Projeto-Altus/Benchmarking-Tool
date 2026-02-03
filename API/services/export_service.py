import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.config import Config

class ExportService:
    
    @staticmethod
    def _clean_currency_value(value):
        if isinstance(value, (int, float)):
            return value
        
        if not isinstance(value, str):
            return value

        clean_val = re.sub(r'[^\d,.-]', '', value).strip()
        
        try:
            if ',' in clean_val:
                clean_val = clean_val.replace('.', '').replace(',', '.')
            return float(clean_val)
        except ValueError:
            return value

    @staticmethod
    def generate_excel(data: list, attributes: list) -> str:
        df = pd.DataFrame(data)
        
        if 'pontuacao_final' in df.columns:
            df['pontuacao_final'] = pd.to_numeric(df['pontuacao_final'], errors='coerce').fillna(0)
            df = df.sort_values(by='pontuacao_final', ascending=False)

        if 'nome_produto' in df.columns:
            df['nome_produto'] = df['nome_produto'].fillna("Produto Sem Nome")
            df['nome_produto'] = df['nome_produto'] + df.groupby('nome_produto').cumcount().astype(str).replace('0', '')
            df.set_index('nome_produto', inplace=True)
        else:
            df.index = [f"Opção {i+1}" for i in range(len(df))]

        fixed_fields = ['pontuacao_final', 'generated_by_ai', 'motivo_escolha', 'url_origem']
        attr_fields = [attr.name for attr in attributes]
        
        for f in fixed_fields + attr_fields:
            if f not in df.columns: df[f] = "N/A"
        
        df = df[fixed_fields + attr_fields]

        def determine_source(row):
            if row.get('generated_by_ai') is True or row.get('generated_by_ai') == 1:
                return "Gerado por IA"
            
            for val in row.values:
                if isinstance(val, str) and "(est.)" in val:
                    return "Dados Reais + IA"
            
            return "Dados Reais"

        df['generated_by_ai'] = df.apply(determine_source, axis=1)

        currency_keywords = ['preço', 'valor', 'custo', 'price']
        for col in df.columns:
            if any(k in col.lower() for k in currency_keywords):
                df[col] = df[col].apply(lambda x: x if (isinstance(x, str) and "(est.)" in x) else ExportService._clean_currency_value(x))

        df_t = df.T
        
        label_map = {
            'pontuacao_final': '🏆 Pontuação (0-100)',
            'generated_by_ai': '🤖 Fonte dos Dados',
            'motivo_escolha': '💡 Análise da IA',
            'url_origem': '🔗 Link da Oferta'
        }
        for attr in attributes:
            label_map[attr.name] = f"{attr.name} (Peso: {attr.importance})"
            
        df_t.rename(index=label_map, inplace=True)
        df_t.reset_index(inplace=True)
        df_t.rename(columns={'index': 'ESPECIFICAÇÕES'}, inplace=True)
        
        if not os.path.exists(Config.DOWNLOAD_FOLDER):
            os.makedirs(Config.DOWNLOAD_FOLDER, exist_ok=True)

        filename = f"comparativo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(Config.DOWNLOAD_FOLDER, filename)
        df_t.to_excel(filepath, index=False, engine='openpyxl')

        wb = load_workbook(filepath)
        ws = wb.active

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        label_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") 
        winner_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") 
        
        ai_full_column_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        estimate_cell_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        estimate_font = Font(color="B45F04", italic=True, bold=True)
        
        header_font = Font(color="FFFFFF", bold=True, size=11)
        bold_font = Font(bold=True)
        link_font = Font(color="0563C1", underline="single")
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 45

        ai_columns_indices = []
        
        for row in ws.iter_rows():
            if "Fonte dos Dados" in str(row[0].value):
                for cell in row:
                    if cell.value == 'Gerado por IA':
                        ai_columns_indices.append(cell.column)
                break 

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align 
                
                if cell.column in ai_columns_indices and cell.column != 1 and cell.row != 1:
                    cell.fill = ai_full_column_fill

                row_label = str(ws.cell(row=cell.row, column=1).value).lower()
                
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    continue

                if cell.column == 1:
                    cell.fill = label_fill
                    cell.font = bold_font
                    cell.alignment = left_align
                    continue

                if cell.column == 2 and cell.column not in ai_columns_indices:
                    cell.fill = winner_fill
                
                if cell.column == 2 and "Pontuação" in row_label:
                    cell.font = bold_font

                val_str = str(cell.value)
                
                if "(est.)" in val_str:
                    clean_text = val_str.replace("(est.)", "").strip()
                    
                    numeric_val = ExportService._clean_currency_value(clean_text)
                    cell.value = numeric_val
                    
                    cell.fill = estimate_cell_fill
                    cell.font = estimate_font

                if "link" in row_label and val_str and "http" in val_str:
                    cell.hyperlink = val_str
                    cell.value = "Ver Oferta 🔗"
                    cell.font = link_font
                
                elif isinstance(cell.value, (int, float)) and ("preço" in row_label or "valor" in row_label):
                    cell.number_format = 'R$ #,##0.00'

                elif "análise" in row_label or "motivo" in row_label:
                    cell.alignment = top_left_align
                    text_len = len(str(cell.value))
                    estimated_height = max(100, (text_len / 50) * 15)
                    current_height = ws.row_dimensions[cell.row].height or 0
                    if estimated_height > current_height:
                        ws.row_dimensions[cell.row].height = min(estimated_height, 300)

                elif "fonte dos dados" in row_label:
                    if "Gerado por IA" == str(cell.value):
                        cell.font = Font(color="9C5700", bold=True)
                    elif "Dados Reais + IA" == str(cell.value):
                        cell.font = Font(color="2F5597", bold=True)
                    else:
                        cell.font = Font(color="006100", bold=True)
                    
                    cell.alignment = center_align

                elif len(str(cell.value)) > 40:
                    cell.alignment = left_align

        wb.save(filepath)
        return filename